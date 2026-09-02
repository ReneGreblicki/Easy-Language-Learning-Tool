from __future__ import annotations

import asyncio
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook

from easy_language_learning_tool.domain.enums import CefrLevel, CefrMode, Language
from easy_language_learning_tool.domain.models import CefrSelection, GenerationSettings, WordRecord
from easy_language_learning_tool.domain.planner import build_generation_plan
from easy_language_learning_tool.generation.service import GenerationService
from easy_language_learning_tool.providers.base import (
    ModelInfo,
    ProviderAdapter,
    ProviderResponse,
    TokenUsage,
)
from easy_language_learning_tool.workbook.service import (
    SENTENCE_HEADERS,
    export_xlsx,
    import_ranked_xlsx,
    import_xlsx,
)


class FakeProvider(ProviderAdapter):
    provider_name = "Fake"

    async def list_models(self) -> list[ModelInfo]:
        return [ModelInfo(id="fake", display_name="Fake")]

    async def generate(self, model: str, prompt: str, schema: dict[str, Any]) -> ProviderResponse:
        marker = "TASKS="
        import json

        tasks = json.loads(prompt[prompt.index(marker) + len(marker) :])
        rows = []
        for task in tasks:
            question = task["sentence_kind"] == "question"
            is_sein = task["lemma"] == "sein"
            sentence = (
                ("Bin ich hier?" if question else "Ich bin hier.")
                if is_sein
                else ("Habe ich Zeit?" if question else "Ich habe Zeit.")
            )
            rows.append(
                {
                    "row_number": task["row_number"],
                    "foreign_word": task["lemma"],
                    "word_translation": task["baseline_translation"] or "translation",
                    "foreign_sentence": sentence,
                    "sentence_translation": "Am I here?" if question else "I am here.",
                    "used_word_form": "bin" if is_sein else "habe",
                    "word_form_or_variant": "present",
                }
            )
        return ProviderResponse(
            rows=rows,
            usage=TokenUsage(input_tokens=20, output_tokens=30),
            raw_model=model,
        )


class GenerationWorkbookTests(unittest.TestCase):
    def test_generation_checkpoint_and_workbook_roundtrip(self) -> None:
        settings = GenerationSettings(
            learning_language=Language.GERMAN,
            translation_language=Language.US_ENGLISH,
            base_sentences=2,
            extra_forms=1,
            question_percentage=Decimal("50"),
            pronoun_change=5,
            cefr=CefrSelection(mode=CefrMode.SINGLE, single_level=CefrLevel.A1),
            seed=7,
        )
        plan = build_generation_plan(
            settings,
            [
                WordRecord(rank=1, lemma="sein", part_of_speech="verb", translation="to be"),
                WordRecord(rank=2, lemma="haben", part_of_speech="verb", translation="to have"),
            ],
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            checkpoint = asyncio.run(
                GenerationService(FakeProvider()).generate(
                    settings=settings,
                    plan=plan,
                    model="fake",
                    checkpoint_path=root / "checkpoint.json",
                    batch_size=2,
                )
            )
            self.assertEqual(len(checkpoint.completed_rows), 4)
            output = root / "sentences.xlsx"
            export_xlsx(
                output,
                checkpoint.completed_rows,
                plan,
                settings,
                provider="Fake",
                model="fake",
                input_tokens=checkpoint.input_tokens,
                output_tokens=checkpoint.output_tokens,
            )
            imported = import_xlsx(output)
            self.assertEqual(len(imported), 4)
            ranked = import_ranked_xlsx(output)
            self.assertEqual([row.rank for row in ranked], [1, 2, 3, 4])
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Sentences", "Metadata"])
            self.assertEqual(
                tuple(cell.value for cell in next(workbook["Sentences"].iter_rows())),
                SENTENCE_HEADERS,
            )
            self.assertEqual(workbook["Sentences"].auto_filter.ref, "A1:D5")
            workbook.close()
            with ZipFile(output) as archive:
                self.assertFalse(
                    any(name.startswith("xl/tables/") for name in archive.namelist()),
                    "Excel table XML must not be emitted; worksheet AutoFilter provides filtering.",
                )


if __name__ == "__main__":
    unittest.main()
