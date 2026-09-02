from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field

from easy_language_learning_tool.domain.enums import Language
from easy_language_learning_tool.domain.models import GenerationSettings, PlannedRow
from easy_language_learning_tool.validation.sentences import GeneratedSentence

SENTENCE_HEADERS = (
    "Foreign-language word",
    "Word translation",
    "Foreign-language sentence",
    "Sentence translation",
)
LEGACY_HEADERS = ("Verb", "Translation", "German Sentence", "English Sentence")
PREVIOUS_HEADERS = (
    "Foreign-language verb",
    "Verb translation",
    "Foreign-language sentence",
    "Sentence translation",
)
METADATA_HEADERS = (
    "Row number",
    "Base item",
    "CEFR level",
    "Frequency rank",
    "Part of speech",
    "Grammatical person",
    "Word form/variant",
    "Question/statement",
    "Generation timestamp",
    "Provider",
    "Model",
    "Validation status",
    "Random seed",
    "Input tokens",
    "Output tokens",
    "Estimated cost USD",
    "Actual cost USD",
    "Generation settings",
)


class WorkbookRow(BaseModel):
    model_config = ConfigDict(frozen=True)

    foreign_word: str = Field(min_length=1)
    word_translation: str = Field(min_length=1)
    foreign_sentence: str = Field(min_length=1)
    sentence_translation: str = Field(min_length=1)


class RankedWorkbookRow(WorkbookRow):
    """One visible workbook data row with its stable, header-free rank."""

    rank: int = Field(ge=1, le=5_000)


def _public_values(row: WorkbookRow) -> tuple[str, str, str, str]:
    return (
        row.foreign_word,
        row.word_translation,
        row.foreign_sentence,
        row.sentence_translation,
    )


def export_xlsx(
    path: Path,
    generated: list[GeneratedSentence],
    plan: tuple[PlannedRow, ...],
    settings: GenerationSettings,
    *,
    provider: str,
    model: str,
    input_tokens: int = 0,
    output_tokens: int = 0,
    estimated_cost_usd: str = "Unknown",
    actual_cost_usd: str = "Unknown",
) -> None:
    if len(generated) != len(plan):
        raise ValueError("Generated rows and plan rows must reconcile exactly.")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sentences = workbook.active
    sentences.title = "Sentences"
    sentences.append(SENTENCE_HEADERS)
    metadata = workbook.create_sheet("Metadata")
    metadata.append(METADATA_HEADERS)
    timestamp = datetime.now(UTC).isoformat()
    settings_json = settings.model_dump_json()
    for sentence, planned in zip(generated, plan, strict=True):
        sentences.append(
            (
                sentence.foreign_word,
                sentence.word_translation,
                sentence.foreign_sentence,
                sentence.sentence_translation,
            )
        )
        metadata.append(
            (
                planned.row_number,
                planned.base_index,
                planned.cefr_level.value,
                planned.word.rank,
                planned.word.part_of_speech,
                planned.grammatical_person.value,
                sentence.word_form_or_variant,
                planned.sentence_kind.value,
                timestamp,
                provider,
                model,
                "accepted",
                settings.seed,
                input_tokens,
                output_tokens,
                estimated_cost_usd,
                actual_cost_usd,
                settings_json,
            )
        )
    for sheet in (sentences, metadata):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E74B5")
    sentences.column_dimensions["A"].width = 24
    sentences.column_dimensions["B"].width = 24
    sentences.column_dimensions["C"].width = 60
    sentences.column_dimensions["D"].width = 60
    workbook.save(path)


def export_csv(path: Path, rows: list[WorkbookRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(SENTENCE_HEADERS)
        writer.writerows(_public_values(row) for row in rows)


def import_ranked_xlsx(path: Path, *, maximum_rows: int = 5_000) -> list[RankedWorkbookRow]:
    if path.suffix.casefold() != ".xlsx":
        raise ValueError("Only .xlsx workbooks are supported.")
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheet = workbook["Sentences"] if "Sentences" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = tuple(str(value).strip() if value is not None else "" for value in next(iterator, ()))
    if headers[:4] not in {SENTENCE_HEADERS, PREVIOUS_HEADERS, LEGACY_HEADERS}:
        workbook.close()
        raise ValueError("Workbook headers do not match the required four-column schema.")
    rows: list[RankedWorkbookRow] = []
    for values in iterator:
        first_four = tuple("" if value is None else str(value).strip() for value in values[:4])
        if not any(first_four):
            continue
        if len(first_four) != 4 or not all(first_four):
            workbook.close()
            raise ValueError(f"Workbook row {len(rows) + 2} contains an empty required cell.")
        rows.append(
            RankedWorkbookRow(
                rank=len(rows) + 1,
                foreign_word=first_four[0],
                word_translation=first_four[1],
                foreign_sentence=first_four[2],
                sentence_translation=first_four[3],
            )
        )
        if len(rows) > maximum_rows:
            workbook.close()
            raise ValueError(f"Workbook exceeds the {maximum_rows:,}-row limit.")
    workbook.close()
    if not rows:
        raise ValueError("Workbook does not contain any sentence rows.")
    return rows


def import_xlsx(path: Path, *, maximum_rows: int = 5_000) -> list[WorkbookRow]:
    """Import public workbook fields while retaining the historical return type."""

    return [
        WorkbookRow.model_validate(row.model_dump(exclude={"rank"}))
        for row in import_ranked_xlsx(path, maximum_rows=maximum_rows)
    ]


def import_language_pair(path: Path) -> tuple[Language, Language] | None:
    """Read the app's language pair from workbook metadata when it is available."""

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "Metadata" not in workbook.sheetnames:
            return None
        sheet = workbook["Metadata"]
        iterator = sheet.iter_rows(values_only=True)
        headers = tuple(
            str(value).strip() if value is not None else "" for value in next(iterator, ())
        )
        try:
            settings_column = headers.index("Generation settings")
        except ValueError:
            return None
        for values in iterator:
            if settings_column >= len(values) or not values[settings_column]:
                continue
            try:
                settings = GenerationSettings.model_validate_json(str(values[settings_column]))
            except (TypeError, ValueError):
                return None
            return settings.learning_language, settings.translation_language
        return None
    finally:
        workbook.close()
